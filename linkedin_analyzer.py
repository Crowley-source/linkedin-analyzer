"""
LinkedIn Profile Analysis Tool
================================
Analyzes and ranks candidate profiles against a job description
using NLP techniques (TF-IDF + cosine similarity).

Designed for Data Analyst workflows: load a CSV of profiles,
score each candidate against a JD, and output a ranked shortlist.

Author  : Data Analyst Portfolio Project
Stack   : Python, Pandas, Scikit-learn
"""

import re
import os
import argparse
import pandas as pd
from sklearn.feature_extraction.text import TfidfVectorizer
from sklearn.metrics.pairwise import cosine_similarity


# ════════════════════════════════════════════════════════════
#  1. TEXT PREPROCESSOR
# ════════════════════════════════════════════════════════════

STOPWORDS = {
    "a","an","the","and","or","but","in","on","at","to","for","of","with",
    "is","are","was","were","be","been","have","has","had","do","does","did",
    "will","would","could","should","may","might","i","we","you","they","it",
    "this","that","these","those","as","by","from","up","about","into","through",
    "during","including","until","while","per","also","its","our","your","their",
    "us","more","than","then","when","where","which","who","how","all","any",
    "both","each","few","most","other","some","such","no","not","only","same",
    "so","too","very","just","can","need","role","experience","looking","work"
}

def clean_text(text: str) -> str:
    if not isinstance(text, str) or not text.strip():
        return ""
    text = text.lower()
    text = re.sub(r"[^a-z0-9\s]", " ", text)
    tokens = [t for t in text.split() if t not in STOPWORDS and len(t) > 2]
    return " ".join(tokens)

def build_profile_text(row: pd.Series) -> str:
    fields = ["headline", "about", "skills", "experience"]
    parts  = [str(row.get(f, "")) for f in fields]
    return " ".join(p for p in parts if p and p != "nan")


# ════════════════════════════════════════════════════════════
#  2. ANALYZER & RANKER
# ════════════════════════════════════════════════════════════

class ProfileAnalyzer:
    """
    Score and rank candidate profiles against a job description
    using TF-IDF vectorization + cosine similarity.
    """

    def __init__(self, ngram_range=(1, 2), max_features=5000):
        self.vectorizer = TfidfVectorizer(
            ngram_range=ngram_range,
            max_features=max_features,
            sublinear_tf=True
        )

    def analyze(self, profiles_df: pd.DataFrame, job_description: str) -> pd.DataFrame:
        df = profiles_df.copy()

        print("── Building profile text ──")
        df["profile_text"]       = df.apply(build_profile_text, axis=1)
        df["profile_text_clean"] = df["profile_text"].apply(clean_text)
        clean_jd                 = clean_text(job_description)

        valid = df["profile_text_clean"].str.len() > 10
        if not valid.all():
            print(f"  ⚠ {(~valid).sum()} profile(s) had insufficient text and scored 0")

        print("── Vectorizing with TF-IDF ──")
        corpus    = [clean_jd] + df.loc[valid, "profile_text_clean"].tolist()
        tfidf_mat = self.vectorizer.fit_transform(corpus)
        scores    = cosine_similarity(tfidf_mat[0], tfidf_mat[1:]).flatten()

        df["raw_score"] = 0.0
        df.loc[valid, "raw_score"] = scores

        max_s = df["raw_score"].max()
        df["match_score"] = (df["raw_score"] / max_s * 100).round(1) if max_s > 0 else 0.0

        df = df.sort_values("match_score", ascending=False).reset_index(drop=True)
        df["rank"]           = df.index + 1
        df["recommendation"] = df["match_score"].apply(self._tier)

        print(f"✓ Scored {len(df)} profiles\n")
        return df

    def top_keywords(self, job_description: str, n: int = 12) -> list:
        clean_jd   = clean_text(job_description)
        vec        = TfidfVectorizer(ngram_range=(1, 2), max_features=2000)
        matrix     = vec.fit_transform([clean_jd])
        names      = vec.get_feature_names_out()
        scores_arr = matrix.toarray().flatten()
        top_idx    = scores_arr.argsort()[::-1][:n]
        return [names[i] for i in top_idx]

    @staticmethod
    def _tier(score: float) -> str:
        if score >= 75:   return "Strong Match"
        elif score >= 50: return "Potential Match"
        elif score >= 25: return "Weak Match"
        else:             return "Not Recommended"


# ════════════════════════════════════════════════════════════
#  3. REPORTER
# ════════════════════════════════════════════════════════════

OUTPUT_COLS = ["rank", "name", "headline", "match_score", "recommendation", "url"]
TIER_EMOJI  = {
    "Strong Match":    "✅",
    "Potential Match": "🟡",
    "Weak Match":      "🟠",
    "Not Recommended": "❌",
}

def print_summary(df: pd.DataFrame, top_n: int = 10):
    print("═" * 66)
    print("  CANDIDATE RANKING REPORT")
    print("═" * 66)
    print(f"  Total profiles analyzed : {len(df)}")
    for tier, emoji in TIER_EMOJI.items():
        print(f"  {emoji} {tier:<22}: {(df['recommendation'] == tier).sum()}")
    print("═" * 66)
    print(f"\n  {'#':<4} {'Name':<26} {'Score':>6}  {'Bar':<22} Tier")
    print("─" * 66)
    for _, row in df.head(top_n).iterrows():
        filled = int(row["match_score"] / 5)
        bar    = "█" * filled + "░" * (20 - filled)
        emoji  = TIER_EMOJI.get(row["recommendation"], "")
        print(f"  #{int(row['rank']):<3} {row['name']:<26} {row['match_score']:>5.1f}%  {bar}  {emoji} {row['recommendation']}")
    print("─" * 66)

def save_csv(df: pd.DataFrame, path: str):
    os.makedirs(os.path.dirname(path), exist_ok=True)
    df[OUTPUT_COLS].to_csv(path, index=False)
    print(f"✓ CSV   → {path}")

def save_excel(df: pd.DataFrame, path: str):
    from openpyxl.styles import PatternFill, Font, Alignment
    os.makedirs(os.path.dirname(path), exist_ok=True)
    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        df[OUTPUT_COLS].to_excel(writer, index=False, sheet_name="Candidates")
        ws = writer.sheets["Candidates"]
        FILLS = {
            "Strong Match":    PatternFill("solid", fgColor="D6F5E3"),
            "Potential Match": PatternFill("solid", fgColor="FFF9C4"),
            "Weak Match":      PatternFill("solid", fgColor="FFE0B2"),
            "Not Recommended": PatternFill("solid", fgColor="FFCDD2"),
        }
        for cell in ws[1]:
            cell.font      = Font(bold=True, color="FFFFFF")
            cell.fill      = PatternFill("solid", fgColor="1B4FFF")
            cell.alignment = Alignment(horizontal="center")
        for row_idx, df_row in enumerate(df[OUTPUT_COLS].itertuples(), start=2):
            fill = FILLS.get(df_row.recommendation)
            for cell in ws[row_idx]:
                if fill:
                    cell.fill = fill
        for col in ws.columns:
            max_len = max(len(str(c.value or "")) for c in col) + 4
            ws.column_dimensions[col[0].column_letter].width = min(max_len, 50)
    print(f"✓ Excel → {path}")


# ════════════════════════════════════════════════════════════
#  4. MAIN
# ════════════════════════════════════════════════════════════

def main(profiles_path: str, job_path: str, output_dir: str = "results"):
    print("\n🔍  LinkedIn Profile Analyzer")
    print("─" * 40)

    print(f"\nLoading profiles : {profiles_path}")
    df = pd.read_csv(profiles_path)
    print(f"✓ {len(df)} profiles loaded")

    print(f"Loading JD       : {job_path}")
    with open(job_path, encoding="utf-8") as f:
        jd = f.read()
    print(f"✓ {len(jd.split())} words\n")

    analyzer  = ProfileAnalyzer()
    ranked_df = analyzer.analyze(df, jd)

    keywords = analyzer.top_keywords(jd)
    print(f"Top JD keywords  : {', '.join(keywords[:10])}\n")

    print_summary(ranked_df)
    save_csv(ranked_df,   os.path.join(output_dir, "ranked_candidates.csv"))
    save_excel(ranked_df, os.path.join(output_dir, "ranked_candidates.xlsx"))
    print("\n✓ Done.")
    return ranked_df


if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="LinkedIn Profile Analyzer")
    parser.add_argument("--profiles", default="data/sample_profiles.csv")
    parser.add_argument("--job",      default="data/job_description.txt")
    parser.add_argument("--output",   default="results")
    args = parser.parse_args()
    main(args.profiles, args.job, args.output)
